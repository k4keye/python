from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 시트 생성
ws.title = "MySheet" # 새로운 시트의 이름변경
ws.sheet_properties.tabColor = "ff77ff" #RGB의 값으로 시트의 색상 변경


ws1 = wb.create_sheet("YouSheet") # 시트 생성과 동시에 이름 지정

ws2 = wb.create_sheet("NewSheet" , 2) # 2번째 인덱스에 시트를 생성

print(wb.sheetnames)
new_ws = wb["NewSheet"]
new_ws.sheet_properties.tabColor = "ff777f"
new_ws["A1"] ="TestA1"

target = wb.copy_worksheet(new_ws)
target.title ="Copied Sheet"


wb.save("test.xlsx")
wb.close()