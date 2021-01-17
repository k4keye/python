from openpyxl import Workbook

wb = Workbook() # 새 워크북 생성 = 엑셀 생성
ws = wb.active
ws.title = "TestSheet"

# A1 셀에 1 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6


print(ws["A1"].value) #A1셀의 정보 출력
print(ws["B2"].value)

print(ws.cell(row=1 , column=1).value)
print(ws.cell(row=2 , column=2).value)


from random import *

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x , column=y , value=randint(0,100))


wb.save("test.xlsx")
wb.close()